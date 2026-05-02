#!/usr/bin/env python3
"""One-time migration: load all rows from jobs.xlsx into Supabase."""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import yaml
from openpyxl import load_workbook
from supabase import create_client


def main():
    config_path = os.environ.get("CONFIG_PATH", "config/config.yaml")
    with open(config_path) as f:
        config = yaml.safe_load(f)

    supabase_cfg = config.get("supabase", {})
    url = supabase_cfg.get("url")
    key = supabase_cfg.get("key")

    if not url or not key:
        print("ERROR: Set supabase.url and supabase.key in config.yaml")
        sys.exit(1)

    client = create_client(url, key)

    excel_file = config.get("storage", {}).get("excel_file", "jobs.xlsx")
    if not os.path.exists(excel_file):
        print(f"ERROR: {excel_file} not found")
        sys.exit(1)

    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append({
            "job_id": str(row[0]) if row[0] else "",
            "requisition_id": str(row[1]) if row[1] else "",
            "title": str(row[2]) if row[2] else "",
            "company": str(row[3]) if row[3] else "",
            "location": str(row[4]) if row[4] else "",
            "date_posted": str(row[5]) if row[5] else "",
            "url": str(row[6]) if row[6] else "",
            "added_on": str(row[7]) if row[7] else None,
        })
    wb.close()

    print(f"Found {len(rows)} jobs in {excel_file}")

    # Insert in batches of 500
    inserted = 0
    skipped = 0
    for i in range(0, len(rows), 500):
        batch = rows[i : i + 500]
        try:
            result = client.table("jobs").upsert(
                batch, on_conflict="title,job_id"
            ).execute()
            inserted += len(result.data)
        except Exception as e:
            print(f"  Batch {i // 500 + 1} error: {e}")
            skipped += len(batch)
        print(f"  Migrated {min(i + 500, len(rows))}/{len(rows)}...")

    print(f"\nDone! Inserted/updated: {inserted}, Skipped/errored: {skipped}")


if __name__ == "__main__":
    main()

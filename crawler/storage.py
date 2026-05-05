import os
import logging
import threading
from datetime import datetime, timezone
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from crawler.parser_base import JobPosting

logger = logging.getLogger(__name__)


try:
    from supabase import create_client, Client as SupabaseClient
    HAS_SUPABASE = True
except ImportError:
    HAS_SUPABASE = False

COLUMNS = [
    "Job ID",
    "Requisition ID",
    "Title",
    "Company",
    "Location",
    "Date Posted",
    "URL",
    "Added On",
]


class ExcelStorage:
    def __init__(self, filepath: str = "jobs.xlsx"):
        self.filepath = filepath
        self._lock = threading.Lock()
        self._ensure_file()

    def _ensure_file(self):
        """Create the Excel file with headers if it doesn't exist."""
        if os.path.exists(self.filepath):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        bold = Font(bold=True)
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold
        wb.save(self.filepath)
        logger.info(f"Created new Excel file: {self.filepath}")

    def get_existing_dedup_keys(self) -> set:
        """Read all title|job_id dedup keys from the spreadsheet."""
        keys = set()
        wb = load_workbook(self.filepath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None and row[2] is not None:
                job_id = str(row[0]).strip().lower()
                title = str(row[2]).strip().lower()
                keys.add(f"{title}|{job_id}")
        wb.close()
        return keys

    def add_jobs(self, jobs: List[JobPosting]) -> List[JobPosting]:
        """Deduplicate and append new jobs. Thread-safe. Returns list of newly added jobs."""
        with self._lock:
            existing_keys = self.get_existing_dedup_keys()
            new_jobs = [j for j in jobs if j.dedup_key not in existing_keys]

            if not new_jobs:
                logger.info("No new jobs to add.")
                return []

            wb = load_workbook(self.filepath)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for job in new_jobs:
                ws.append([
                    job.job_id,
                    job.requisition_id,
                    job.title,
                    job.company,
                    job.location,
                    job.date_posted,
                    job.url,
                    now,
                ])

            wb.save(self.filepath)
            logger.info(f"Added {len(new_jobs)} new jobs to {self.filepath}")
            return new_jobs

    def get_jobs_added_today_count(self) -> int:
        """Count jobs added today by checking the 'Added On' column (index 8)."""
        today_prefix = datetime.now().strftime("%Y-%m-%d")
        count = 0
        with self._lock:
            wb = load_workbook(self.filepath, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[7] is not None and str(row[7]).startswith(today_prefix):
                    count += 1
            wb.close()
        return count


class SupabaseStorage:
    """Storage backend using Supabase (Postgres). Same interface as ExcelStorage."""

    def __init__(self, url: str, key: str):
        if not HAS_SUPABASE:
            raise ImportError("supabase package is required: pip install supabase")
        self.client: SupabaseClient = create_client(url, key)
        self._lock = threading.Lock()

    def get_existing_dedup_keys(self) -> set:
        """Fetch all title|job_id dedup keys from Supabase."""
        keys = set()
        offset = 0
        batch_size = 1000
        while True:
            result = (
                self.client.table("jobs")
                .select("title,job_id")
                .range(offset, offset + batch_size - 1)
                .execute()
            )
            if not result.data:
                break
            for row in result.data:
                title = row["title"].strip().lower()
                job_id = row["job_id"].strip().lower()
                keys.add(f"{title}|{job_id}")
            if len(result.data) < batch_size:
                break
            offset += batch_size
        return keys

    def add_jobs(self, jobs: List[JobPosting]) -> List[JobPosting]:
        """Deduplicate and insert new jobs into Supabase. Returns newly added jobs."""
        with self._lock:
            existing_keys = self.get_existing_dedup_keys()
            new_jobs = [j for j in jobs if j.dedup_key not in existing_keys]

            if not new_jobs:
                logger.info("Supabase: No new jobs to add.")
                return []

            now = datetime.now(timezone.utc).isoformat()
            rows = []
            for job in new_jobs:
                row = {
                    "job_id": job.job_id,
                    "requisition_id": job.requisition_id,
                    "title": job.title,
                    "company": job.company,
                    "location": job.location,
                    "date_posted": job.date_posted,
                    "url": job.url,
                    "added_on": now,
                }
                if job.description:
                    row["description"] = job.description
                rows.append(row)

            # Insert in batches of 500
            for i in range(0, len(rows), 500):
                batch = rows[i : i + 500]
                self.client.table("jobs").insert(batch).execute()

            logger.info(f"Supabase: Added {len(new_jobs)} new jobs")
            return new_jobs

    def get_jobs_added_today_count(self) -> int:
        """Count jobs added today."""
        today_start = datetime.now(timezone.utc).replace(
            hour=0, minute=0, second=0, microsecond=0
        ).isoformat()
        result = (
            self.client.table("jobs")
            .select("id", count="exact")
            .gte("added_on", today_start)
            .execute()
        )
        return result.count or 0


class DualStorage:
    """Writes to both ExcelStorage and SupabaseStorage. Excel is primary."""

    def __init__(self, excel: ExcelStorage, supabase: Optional[SupabaseStorage] = None):
        self.excel = excel
        self.supabase = supabase

    def add_jobs(self, jobs: List[JobPosting]) -> List[JobPosting]:
        new_jobs = self.excel.add_jobs(jobs)
        if self.supabase and new_jobs:
            try:
                self.supabase.add_jobs(jobs)
            except Exception as e:
                logger.error(f"Supabase write failed (Excel succeeded): {e}")
        return new_jobs

    def get_jobs_added_today_count(self) -> int:
        return self.excel.get_jobs_added_today_count()

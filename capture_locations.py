"""
One-time script to capture raw location strings from every enabled company.
Saves results to location_analysis.xlsx with columns:
  Company | Parser | Location (raw) | Count
"""
import yaml
import logging
import sys
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Import all parsers (same as main.py)
from crawler.parsers.salesforce import SalesforceParser
from crawler.parsers.airbnb import AirbnbParser
from crawler.parsers.apple import AppleParser
from crawler.parsers.caterpillar import CaterpillarParser
from crawler.parsers.eightfold import EightfoldParser
from crawler.parsers.workday import WorkdayParser
from crawler.parsers.generic import GenericHTMLParser
from crawler.parsers.spotify import SpotifyParser
from crawler.parsers.oracle_hcm import OracleHCMParser
from crawler.parsers.visa import VisaParser
from crawler.parsers.microsoft import MicrosoftParser
from crawler.parsers.meta import MetaParser
from crawler.parsers.stripe import StripeParser
from crawler.parsers.uber import UberParser
from crawler.parsers.databricks import DatabricksParser
from crawler.parsers.paypal import PayPalParser
from crawler.parsers.ford import FordParser
from crawler.parsers.phenom import PhenomParser
from crawler.parsers.walmart import WalmartParser
from crawler.parsers.jibe import JibeParser
from crawler.parsers.radancy import RadancyParser
from crawler.parsers.greenhouse import GreenhouseParser
from crawler.parsers.greenhouse_api import GreenhouseAPIParser
from crawler.parsers.amazon import AmazonParser
from crawler.parsers.avature import AvatureParser

PARSER_REGISTRY = {
    "salesforce": SalesforceParser,
    "airbnb": AirbnbParser,
    "apple": AppleParser,
    "caterpillar": CaterpillarParser,
    "eightfold": EightfoldParser,
    "workday": WorkdayParser,
    "generic": GenericHTMLParser,
    "spotify": SpotifyParser,
    "oracle_hcm": OracleHCMParser,
    "visa": VisaParser,
    "microsoft": MicrosoftParser,
    "meta": MetaParser,
    "stripe": StripeParser,
    "uber": UberParser,
    "databricks": DatabricksParser,
    "paypal": PayPalParser,
    "ford": FordParser,
    "phenom": PhenomParser,
    "walmart": WalmartParser,
    "jibe": JibeParser,
    "radancy": RadancyParser,
    "greenhouse": GreenhouseParser,
    "greenhouse_api": GreenhouseAPIParser,
    "amazon": AmazonParser,
    "avature": AvatureParser,
}


def fetch_locations(site_config):
    """Fetch jobs from a site and return (company, parser, list_of_locations)."""
    name = site_config["name"]
    parser_name = site_config.get("parser", "")
    parser_cls = PARSER_REGISTRY.get(parser_name)
    if not parser_cls:
        return name, parser_name, [], "Unknown parser"
    try:
        parser = parser_cls(site_config)
        jobs = parser.fetch_and_parse()
        locations = [j.location for j in jobs]
        logger.info(f"  {name} ({parser_name}): {len(jobs)} jobs fetched")
        return name, parser_name, locations, None
    except Exception as e:
        logger.error(f"  {name} ({parser_name}): ERROR - {e}")
        return name, parser_name, [], str(e)


def main():
    with open("config.yaml") as f:
        cfg = yaml.safe_load(f)

    sites = cfg.get("sites", [])
    # Take first enabled entry per company (avoid duplicate fetches)
    seen_companies = set()
    unique_sites = []
    for s in sites:
        if not s.get("enabled", True):
            continue
        company = s["name"]
        if company not in seen_companies:
            seen_companies.add(company)
            unique_sites.append(s)

    logger.info(f"Fetching locations from {len(unique_sites)} companies...")

    # Run in parallel (max 4 threads to be polite)
    results = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(fetch_locations, s): s["name"] for s in unique_sites}
        for future in as_completed(futures):
            result = future.result()
            results.append(result)

    # Sort by company name
    results.sort(key=lambda x: x[0].lower())

    # Build xlsx
    wb = Workbook()

    # --- Sheet 1: Raw locations (every unique location per company) ---
    ws_raw = wb.active
    ws_raw.title = "Raw Locations"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Company", "Parser", "Location (Raw)", "Count", "Error"]
    for col, h in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for company, parser_name, locations, error in results:
        if error and not locations:
            ws_raw.cell(row=row, column=1, value=company).border = thin_border
            ws_raw.cell(row=row, column=2, value=parser_name).border = thin_border
            ws_raw.cell(row=row, column=3, value="").border = thin_border
            ws_raw.cell(row=row, column=4, value=0).border = thin_border
            err_cell = ws_raw.cell(row=row, column=5, value=error)
            err_cell.border = thin_border
            err_cell.font = Font(color="FF0000")
            row += 1
            continue

        loc_counts = Counter(locations)
        for loc, count in sorted(loc_counts.items(), key=lambda x: -x[1]):
            ws_raw.cell(row=row, column=1, value=company).border = thin_border
            ws_raw.cell(row=row, column=2, value=parser_name).border = thin_border
            ws_raw.cell(row=row, column=3, value=loc).border = thin_border
            ws_raw.cell(row=row, column=4, value=count).border = thin_border
            ws_raw.cell(row=row, column=5, value="").border = thin_border
            row += 1

    # Auto-width
    for col in ws_raw.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_raw.column_dimensions[col_letter].width = min(max_len + 3, 80)

    # --- Sheet 2: Summary - country pattern per company ---
    ws_summary = wb.create_sheet("Country Pattern Summary")
    sum_headers = ["Company", "Parser", "Total Jobs", "Sample Locations (top 5)", "Country Format Detected"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for company, parser_name, locations, error in results:
        if error and not locations:
            continue
        total = len(locations)
        loc_counts = Counter(locations)
        top5 = [f"{loc} ({cnt})" for loc, cnt in loc_counts.most_common(5)]

        # Detect country format
        all_locs_lower = " | ".join(locations).lower()
        formats = []
        if "united states of america" in all_locs_lower:
            formats.append("United States of America")
        if "united states" in all_locs_lower and "united states of america" not in all_locs_lower:
            formats.append("United States")
        # Check for standalone "USA" (not part of another word)
        for loc in locations:
            if "USA" in loc.split() or loc.strip().endswith("USA") or ", USA" in loc:
                formats.append("USA")
                break
        for loc in locations:
            parts = loc.replace(",", " ").split()
            if "US" in parts:
                formats.append("US")
                break
        # Check for state abbreviations only (no country)
        import re
        state_pattern = re.compile(r'\b[A-Z]{2}\b')
        has_state_only = False
        for loc in locations[:20]:
            if state_pattern.search(loc) and not any(c in loc.lower() for c in ["united states", "usa", ", us"]):
                has_state_only = True
                break
        if has_state_only and not formats:
            formats.append("City, State only (no country)")
        if not formats:
            if locations:
                formats.append("Other/Unknown")
            else:
                formats.append("No data")

        ws_summary.cell(row=row, column=1, value=company).border = thin_border
        ws_summary.cell(row=row, column=2, value=parser_name).border = thin_border
        ws_summary.cell(row=row, column=3, value=total).border = thin_border
        ws_summary.cell(row=row, column=4, value=" | ".join(top5)).border = thin_border
        ws_summary.cell(row=row, column=5, value=", ".join(formats)).border = thin_border
        row += 1

    for col in ws_summary.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = min(max_len + 3, 100)

    outfile = "location_analysis.xlsx"
    wb.save(outfile)
    logger.info(f"\nSaved to {outfile}")
    logger.info(f"  Sheet 'Raw Locations': every unique location per company")
    logger.info(f"  Sheet 'Country Pattern Summary': country format detected per company")


if __name__ == "__main__":
    main()
